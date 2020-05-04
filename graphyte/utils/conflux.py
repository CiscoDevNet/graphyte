#!/usr/bin/env python3
"""conflux.py

Set of high level operations to interact with a Confluence instance.

Extends atlassian's Confluence.

maintainer: Jorge Somavilla (@Cisco)

requires: atlassian
          xlrd
          openpyxl
          pandas

"""

import json
import requests
import urllib3
import re
import os
import logging
from atlassian import Confluence
import xlrd
import openpyxl
import pandas as pd


log = logging.getLogger(__name__)

class Conflux(Confluence):

    def test_connection(self):
        """
        Return boolean True if connection worked,
        False otherwise.
        :param url: Page URL
        :return: Boolean
        """
        urllib3.disable_warnings()
        headers={'Content-Type':'application/json'}
        user = self.username
        pwd = self.password
        auth=(user, pwd)
        r = requests.get(self.url, verify=False, headers=headers, auth=auth)
        if "<title>HTTP Status 401 – Unauthorized</title>" in r.text:
            return False
        else:
            return True

    def create_empty_page_get_id (self,title,parent_id):
        """
        Creates empty child page under parent page
        and returns child page ID.
        :param filepath: Child page title
        :param parent_id: Confluence ID of parent page
        :return: Child page ID
        """
        space = self.get_page_space(parent_id)
        status = self.create_page(
            space=space,
            title=title,
            body='',
            parent_id=parent_id,
            type='page',
            representation='storage'
        )
        if 'id' in status:
            page_id = (status["id"])
        else:
            log.error('Could not create page and retrieve page id.')
            return
        return page_id


    def attach_file_get_id (self,filepath,page_id,content_type='application/binary'):
        """
        Add file as attachment. Optionally specifying
        content-type (default 'application/binary)'.
        :param filepath: Path to file to be attached
        :param page_id: confluence id of target page
        :return: attachment ID,file name
        """
        filename = os.path.basename(filepath)
        space = self.get_page_space(page_id)
        status = self.attach_file(
            filename=filepath,
            name=filename,
            content_type=content_type,
            page_id=page_id,
            space=space
            )
        if not status:
            log.error('Error uploading file %s: POST operation unsuccessful' % (filename))
            return
        if 'results' in status:
            if 'id' in status['results'][0]:
                r_id = (status['results'][0]['id'])
            else:
                log.error('Error uploading file %s: id not found' % (filename))
                return
        else:
            if 'id' in status:
                r_id = (status['id'])
            else:
                log.error('Error uploading file %s: id not found' % (filename))
                return
        return r_id,filename

    def attach_svg_get_id (self,filepath,page_id):
        """
        Add SVG file as attachment. It is added with
        content-type='image/svg+xml'.
        :param filepath: Path to file to be attached
        :param page_id: confluence id of target page
        :return: attachment ID,file name
        """
        try:
            r_id, filename = self.attach_file_get_id(filepath, page_id, content_type='image/svg+xml')
        except ValueError:
            log.error('Could not attach svg file.')
            return
        return r_id,filename

    def get_page_title(self,page_id):
        """
        Return Confluence ID of page given its URL.
        Accepts both "/display" and "/rest" URLs.
        :param url: Page URL
        :return: Confluence ID
        """
        status = self.get_page_by_id(
            page_id=page_id
        )
        if 'title' in status:
            title = status["title"]
        else:
            log.error('Could not successfully create page and retrieve page id.')
            return
        return title

    def get_page_id(self, url):
        """
        Return Confluence ID of page given its URL.
        Accepts both "/display" and "/rest" URLs.
        :param url: Page URL
        :return: Confluence ID
        """
        urllib3.disable_warnings()
        headers={'Content-Type':'application/json'}
        user = self.username
        pwd = self.password
        auth=(user, pwd)
        r = requests.get(url, verify=False, headers=headers, auth=auth)
        if re.match(r'.*/rest/api/.*', url):
            id = self.get_page_id_from_json(r.json())
        else:
            id = self.get_page_id_from_html(r.text)
        return id

    def get_page_id_from_json(self,data):
        """
        Extract Confluence page ID from page JSON.
        :param data: page JSON
        :return: Confluence ID
        """
        if 'results' in data:
            for results in data["results"]:
                id = (results["id"])
        else:
            id = data["id"]
        return id

    def get_page_id_from_html(self,data):
        """
        Extract Confluence page ID from HTML.
        :param data: HTML code of page
        :return: Confluence ID
        """
        try:
            match = re.search('<meta name=\"ajs-page-id\" content=\"(\d+)\">', data).group(1)
        except AttributeError:
            match = False
        if match:
            id = match
        return id

    def prepend_header_to_page(self, page_id, header, header_type=1):
        """
        Append header field to page body, optionally choosing
        header type (default 1 -> h1)
        :param page_id: confluence id of target page
        :param header: header text
        :param header_type: Level of the header.
                            One of {1|2|3|4|5|6|7}
        :return: True if all ok
        """
        body = "<p><h" + header_type +">" + header + "</h" + header_type + "></p>"
        self.prepend_to_page(page_id, body)
        return True

    def append_header_to_page(self, page_id, header, header_type=1):
        """
        Append header field to page body, optionally choosing
        header type (default 1 -> h1)
        :param page_id: confluence id of target page
        :param header: header text
        :param header_type: Level of the header.
                            One of {1|2|3|4|5|6|7}
        :return: True if all ok
        """
        body = "<p><h" + header_type +">" + header + "</h" + header_type + "></p>"
        self.append_to_page(page_id, body)
        return True

    def append_p_to_page(self, page_id, p):
        """
        Append <p>p</p> element to page body.
        :param page_id: confluence id of target page
        :param p: text to add inside <p> element
        :return: True if all ok
        """
        body = "<p>" + p + "</p>"
        self.append_to_page(page_id, body)
        return True

    def append_to_page(self, page_id, body):
        """
        Append body to page.
        :param page_id: confluence id of target page
        :param body: HTML body to be appended
        :return: True if all ok
        """
        title = self.get_page_title(page_id)
        status = self.append_page(
            title=title,
            page_id=page_id,
            append_body=body
        )
        return True

    def prepend_to_page(self, page_id, body):
        """
        Prepend Table of Contents to existing page
        :param page_id: confluence id of target page
        :param max_level: maximum title depth level
        :return: True if all ok
        """
        title = self.get_page_title(page_id)
        status = self.get_page_by_id(
            page_id=page_id,
            expand='body.storage'
        )
        body = body + status['body']['storage']['value']
        self.update_page(page_id, title, body)
        return True

    def append_file_content_to_page(self, page_id, file):
        """
        Append file content to page body.
        :param page_id: confluence id of target page
        :param file: path to file to be appended
        :return: True if all ok
        """
        with open(file, "r") as f:
            body = f.read()
        self.append_p_to_page(page_id, body)
        return True

    def urlify_name(self,filepath):
        """
        Return file name with '%20' instead of spaces.
        :param filepath: path to file
        :return: name of file in URL format
        """
        filename = os.path.basename(filepath)
        urlified = re.sub(r' ', r'%20', filename.rstrip())
        return urlified

    def attach_svg_append_as_img(self, page_id, svg_file):
        """
        Upload SVG file as attachment to page, and append
        to page body as HTML <img> tag.
        :param page_id: confluence id of target page
        :param svg_file: path to SVG file to be attached
        :return: True if all ok
        """
        att_id = self.attach_svg_get_id(svg_file,page_id)
        body = "<p><span class=\"confluence-embedded-file-wrapper\">" \
               "<img class=\"confluence-embedded-image\" " \
               "src=\"/conf/download/attachments/" + page_id + "/" \
               + self.urlify_name(svg_file) \
               + "\" data-image-src=\"/conf/download/attachments/" \
               + page_id + "/" + self.urlify_name(svg_file) \
               + "\" data-unresolved-comment-count=\"0\" data-linked-resource-id=\"" \
               + str(att_id) + "\" data-linked-resource-version=\"3\" " \
               + "data-linked-resource-type=\"attachment\" " \
                 "data-linked-resource-default-alias=\"" + svg_file \
               + "\" data-base-url=\"" + self.url \
               + "\" data-linked-resource-content-type=\"image/svg+xml\" " \
                 "data-linked-resource-container-id=\"" + page_id \
               + "\" data-linked-resource-container-version=\"63\"></img></span></p>"
        title = self.get_page_title(page_id)
        status = self.append_page(
            title=title,
            page_id=page_id,
            append_body=body
        )
        return True

    def append_body_to_page(self, page_id, body):
        """
        Append HTML body to existing page.
        :param page_id: confluence id of target page
        :param body: HTML body snippet to be added to page
        :return: True if all ok
        """
        title = self.get_page_title(page_id)
        status = self.append_page(
            title=title,
            page_id=page_id,
            append_body=body
        )
        return True

    def build_attachchment_href(self,page_id,att_name,text):
        """
        Build HTML href for existing attachment.
        :param page_id: confluence id of target page
        :param att_name: filename of attachment
        :param text: text to display in the href
        :return: the HTML href snippet
        """
        return "<a href=\"" + self.url + "/download/attachments/" \
               + page_id + "/" + att_name + "?api=v2\" rel=\"nofollow\">" \
               + text + "</a>"

    def remove_children_pages(self, page_id):
        """
        Remove children pages of given page.
        :param page_id: confluence id of target page
        :return: True if all ok
        """
        children_pages = self.get_page_child_by_type(page_id)
        for children_page in children_pages:
            self.remove_page(children_page.get('id'), "", False)
        return True

    def build_template_body(self,filepath):
        """
        Return HTML with text file contents inside single cell table.
        :param filepath: text file to be dumped into HTML body
        :return: HTML table body with file contents
        """
        with open(filepath, encoding="utf8", errors="ignore") as tf:
            body = "<table class=\"wrapped relative-table\" style=\"width: 100.0%;\"><colgroup>" \
                   "<col style=\"width: 100.0%;\" /></colgroup><tbody><tr><th><div class=\"content-wrapper\">"
            for l in tf:
                l = re.sub(r'&', r'&amp;', l.rstrip())
                l = re.sub(r'<', r'&lt;', l.rstrip())
                l = re.sub(r'>', r'&gt;', l.rstrip())
                if l == "":
                    l = "<br/>"
                body = body + "<pre>" + l + "</pre>"
        body = body + "</div></th></tr></tbody></table>"
        return body


    def download_all_attachments(self, page_id, dir, pattern=".*"):
        """
        Download all attachments in page, optionally matching regex pattern
        to filenames.
        :param page_id: confluence id of target page
        :param dir: target directory where attachments should be downloaded
        :param pattern: regex pattern to filter filenames to be downloaded
        :return: True if all ok
        """
        d = self.get_attachments_urls(page_id)
        for name in d:
            if re.match(pattern, name):
                auth = (self.username, self.password)
                r = requests.get(self.url + d[name], verify=False, auth=auth)
                with open(dir + '/' + name, 'wb') as f:
                    f.write(r.content)
        return True


    def get_attachments_urls(self, page_id, jump=50):
        """
        Get download URLs of all attachments
        :param page_id: confluence id of target page
        :param jump: pagination size when listing attachments, optional.
        :return: dictionary of filename:download_url
        """
        d = dict()
        params = {}
        params['start'] = 0
        params['limit'] = jump
        next = 'rest/api/content/{id}/child/attachment'\
            .format(id=page_id, params=params)
        while next:
            result = self.get(next)
            for att in result['results']:
                title = att['title']
                dl = att['_links']['download']
                d[title.rstrip()] = dl.rstrip()
            if 'next' in result['_links']:
                next = result['_links']['next']
            else:
                next = ''
        return d


    def prepend_toc_to_page(self, page_id, max_level=7):
        """
        Prepend Table of Contents to existing page
        :param page_id: confluence id of target page
        :param max_level: maximum title depth level
        :return: True if all ok
        """
        title = self.get_page_title(page_id)
        status = self.get_page_by_id(
            page_id=page_id,
            expand='body.storage'
        )
        toc = "<p><ac:structured-macro ac:name=\"toc\" " \
              "ac:schema-version=\"1\"><ac:parameter " \
              "ac:name=\"maxLevel\">" + str(max_level) + \
              "</ac:parameter></ac:structured-macro></p>"
        toc = self.build_scroll_ignore(toc)
        body = toc + status['body']['storage']['value']
        self.update_page(page_id, title, body)
        return True

    def build_toc(self, max_level=7):
        """
        Prepend Table of Contents to existing page
        :param max_level: maximum title depth level
        :return: toc body
        """
        body = "<p><ac:structured-macro ac:name=\"toc\" " \
              "ac:schema-version=\"1\"><ac:parameter " \
              "ac:name=\"maxLevel\">" + str(max_level) + \
              "</ac:parameter><ac:parameter ac:name=\"exclude\">" \
              "Table of Contents</ac:parameter>" \
              "</ac:structured-macro></p>"
        return body

    def build_toc_with_header(self, header, header_level="1", max_level=7, ignore_toc_header=True):
        """
        Prepend Table of Contents to existing page
        :param max_level: maximum title depth level
        :return: toc body
        """
        body = "<h" + header_level + ">" + header + "</h" + header_level + "><p><ac:structured-macro ac:name=\"toc\" " \
              "ac:schema-version=\"1\"><ac:parameter " \
              "ac:name=\"maxLevel\">" + str(max_level)
        if ignore_toc_header:
            body = body + "</ac:parameter><ac:parameter ac:name=\"exclude\">" + \
                  header + "</ac:parameter>"
        body = body + "</ac:structured-macro></p>"
        return body

    def pprint(self,data):
        '''
        Serializes to JSON and pretty prints data
        received as a text/binary python object.
        :param data: python object containing data.
        :return:
        '''
        print (json.dumps(
               data,
               sort_keys = True,
               indent = 4,
               separators = (', ', ' : '))
               )

    def append_workbook_as_tables(self, page_id, workbook, sheet_name=None):
        '''
        Appends workbook sheets as tables to target page body.
        Optionally select title of sheet to be appended, by default
        all sheets in workbook are added to the page.
        :param page_id: confluence id of target page
        :param workbook: excel workbook file.
        :param sheet_name: optional, name of sheet to process
        (will ignore the others). By default processes all sheets.
        :return: True if all ok
        '''
        # todo: if xls call xls_to_xlsx ()
        wb = openpyxl.load_workbook(workbook)
        if sheet_name:
            match = False
            for sn in wb.sheetnames:
                if sn == sheet_name:
                    match = True
            if match:
                s = wb[sheet_name]
                self.append_sheet_as_table(page_id,s)
            else:
                log.error('Error fetching sheet %s: '
                          'not found in workbook %s'
                          % (sheet_name,workbook))
                return False
        else:
            sheet_names = wb.get_sheet_names()
            for sn in sheet_names:
                s = wb[sn]
                self.append_sheet_as_table(page_id,s)
        return True

    def append_sheet_as_table(self, page_id, sheet):
        '''
        Adds contents of openpyxl sheet to target page
        as an HTML table inside a scroll-title element.
        :param page_id: confluence id of target page
        :param sheet: openpyxl sheet object.
        :return: True if all ok
        '''
        body = "<p class=\"auto-cursor-target\"><br /></p>" \
               "<ac:structured-macro ac:name=\"scroll-title\" ac:schema-version=\"1\">" \
               "<ac:parameter ac:name=\"title\">" + sheet.title + \
               "</ac:parameter><ac:rich-text-body>" \
               "<p class=\"auto-cursor-target\"><br /></p>"
        body = body + "<table><tbody>"
        for row in sheet.iter_rows(max_row=sheet.max_row, max_col=sheet.max_column):
            body = body + "<tr>"
            for cell in row:
                v = str(cell.value) or ""
                v = re.sub(r'<', r'&lt;', v)
                v = re.sub(r'>', r'&gt;', v)
                v = re.sub(r'"', r'&quot;', v)
                v = re.sub(r'\\', r'\\\\', v)
                if str(cell.row) == "1":
                    body = body + "<th>" + v + "</th>"
                else:
                    body = body + "<td>" + v + "</td>"
            body = body + "</tr>"
        body = body + "</tbody></table>"
        body = body + "<p class=\"auto-cursor-target\"><br /></p>" \
                      "</ac:rich-text-body></ac:structured-macro><p><br /></p>"
        self.append_body_to_page(page_id,body)
        return True

    def append_csv_as_table(self, page_id, csv_file):
        '''
        Adds contents of csv file to target page
        as an HTML table inside a scroll-title element.
        :param page_id: confluence id of target page
        :param csv_file: csv file.
        :return: True if all ok
        '''
        c = pd.read_csv(csv_file)
        table = c.to_html()
        body = "<p class=\"auto-cursor-target\"><br /></p>" \
               "<ac:structured-macro ac:name=\"scroll-title\" ac:schema-version=\"1\">" \
               "<ac:parameter ac:name=\"title\">" + os.path.basename(csv_file) + \
               "</ac:parameter><ac:rich-text-body>" \
               "<p class=\"auto-cursor-target\"><br /></p>" + table + \
               "<p class=\"auto-cursor-target\"><br /></p>" \
               "</ac:rich-text-body></ac:structured-macro><p><br /></p>"
        self.append_body_to_page(page_id, body)
        return True

    def append_children_macro(self, page_id, depth="1"):
        '''
        Adds a list of child pages to page body.
        :param page_id: confluence id of target page
        :param depth: number of child levels to display.
        :return: True if all ok
        '''
        body = "<p><ac:structured-macro ac:name=\"children\" " \
               "ac:schema-version=\"2\">" \
               "<ac:parameter ac:name=\"depth\">" + depth + \
               "</ac:parameter><ac:parameter ac:name=\"sort\">" \
               "creation</ac:parameter></ac:structured-macro></p>"
        self.append_body_to_page(page_id,body)
        return True

    def xls_to_xlsx(self, filename):
        '''
        Support legacy XLS files. Using
        XLRD (unmaintained) to convert XLS to
        XLSX.
        :param filename: xls file.
        :return: xlsx file
        '''
        book = xlrd.open_workbook(filename)
        index = 0
        nrows, ncols = 0, 0
        while nrows * ncols == 0:
            sheet = book.sheet_by_index(index)
            nrows = sheet.nrows
            ncols = sheet.ncols
            index += 1

        # prepare an openpyxl xlsx sheet
        book1 = Workbook()
        sheet1 = book1.get_active_sheet()

        for row in xrange(0, nrows):
            for col in xrange(0, ncols):
                sheet1.cell(row=row, column=col).value = sheet.cell_value(row, col)

        return book1

    def build_scroll_ignore(self, body):
        '''
        Adds contents of openpyxl sheet to target page
        as an HTML table inside a scroll-title element.
        :param sheet: openpyxl sheet object.
        :return: html body inside scroll element
        '''
        body = "<ac:structured-macro ac:name=\"scroll-ignore\" ac:schema-version=\"1\" >" \
            "<ac:parameter ac:name=\"scroll-pdf\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-office\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-chm\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-docbook\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-eclipsehelp\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-epub\">true</ac:parameter>" \
            "<ac:parameter ac:name=\"scroll-html\">true</ac:parameter>" \
            "<ac:rich-text-body>" + body + "</ac:rich-text-body></ac:structured-macro>"
        return body